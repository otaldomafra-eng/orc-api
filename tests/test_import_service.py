from pathlib import Path
from decimal import Decimal

from openpyxl import Workbook
from sqlalchemy import create_engine, select
from sqlalchemy.orm import Session

from app.caixa.parser import parse_sinapi_package
from app.db import Base
from app.models import SinapiCompetencia, SinapiItem
from app.services.imports import publish_parsed_file


def create_minimal_reference_xlsx(path: Path) -> None:
    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)

    insumos = workbook.create_sheet("ISE")
    insumos["A3"] = "Mês de Referência:"
    insumos["B3"] = "04/2026"
    insumos.append([])
    insumos.append([])
    insumos.append([])
    insumos.append([])
    insumos.append([])
    insumos.append([])
    insumos.append(["Classificacao", "Codigo do Insumo", "Descricao do Insumo", "Unidade", "Origem de Preco", "AC", "TO"])
    insumos.append(["MATERIAL", "1379", "CIMENTO PORTLAND", "KG", "C", 1.00, 1.04])

    composicoes = workbook.create_sheet("CSE")
    composicoes["A3"] = "Mês de Referência:"
    composicoes["B3"] = "04/2026"
    for _ in range(5):
        composicoes.append([])
    composicoes.append([None, None, None, None, "AC", None, "TO", None])
    composicoes.append(["Grupo", "Codigo da Composicao", "Descricao", "Unidade", "Custo (R$)", "%AS", "Custo (R$)", "%AS"])
    composicoes.append(["Pisos", "87248", "ARGAMASSA", "M3", 500, 0, 620, 0.12])

    workbook.save(path)


def test_parse_and_publish_minimal_consolidated_xlsx(tmp_path: Path) -> None:
    xlsx_path = tmp_path / "SINAPI_Referencia_2026_04.xlsx"
    create_minimal_reference_xlsx(xlsx_path)

    parsed = parse_sinapi_package(xlsx_path, tmp_path / "extract")

    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    with Session(engine) as session:
        competencia = publish_parsed_file(
            session,
            parsed,
            source_url="manual://fixture",
            source_filename=xlsx_path.name,
            file_sha256="b" * 64,
        )
        session.commit()
        items = session.scalars(select(SinapiItem).order_by(SinapiItem.codigo)).all()
        saved_competencia = session.get(SinapiCompetencia, competencia.id)

    assert saved_competencia is not None
    assert saved_competencia.uf == "TO"
    assert [item.codigo for item in items] == ["1379", "87248"]
    assert items[0].valor_sem_encargos == Decimal("1.040000")
