from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook

from app.errors import AppError


@dataclass(frozen=True)
class ParsedItem:
    codigo: str
    descricao: str
    tipo: str
    unidade: str | None
    classe: str | None
    valor_sem_encargos: Decimal | None
    valor_onerado: Decimal | None
    valor_nao_onerado: Decimal | None
    raw_data: dict


@dataclass(frozen=True)
class ParsedComposicaoItem:
    composicao_codigo: str
    item_codigo: str
    item_descricao: str
    item_tipo: str | None
    unidade: str | None
    coeficiente: Decimal | None
    valor_sem_encargos: Decimal | None
    valor_onerado: Decimal | None
    valor_nao_onerado: Decimal | None
    raw_data: dict


@dataclass(frozen=True)
class ParsedSinapiFile:
    uf: str
    ano: int
    mes: int
    source_layout: str
    items: list[ParsedItem]
    composicao_items: list[ParsedComposicaoItem]


def parse_decimal(value) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, int | float):
        return Decimal(str(value))
    text = str(value).strip()
    if text in {"-", "–"}:
        return None
    text = text.replace(".", "").replace(",", ".")
    try:
        return Decimal(text)
    except Exception as exc:
        raise AppError("import_validation_failed", "Valor decimal invalido", 400, {"value": str(value)}) from exc


def extract_xlsx_files(path: Path, target_dir: Path) -> list[Path]:
    if path.suffix.lower() == ".xlsx":
        return [path]
    if path.suffix.lower() != ".zip":
        raise AppError("unsupported_file_layout", "Arquivo deve ser ZIP ou XLSX", 400)

    target_dir.mkdir(parents=True, exist_ok=True)
    xlsx_files: list[Path] = []
    with ZipFile(path) as archive:
        for member in archive.namelist():
            member_path = Path(member)
            if member_path.is_absolute() or ".." in member_path.parts:
                raise AppError("unsupported_file_layout", "ZIP contem caminho inseguro", 400)
            if member.lower().endswith(".xlsx"):
                archive.extract(member, target_dir)
                xlsx_files.append(target_dir / member)
    if not xlsx_files:
        raise AppError("unsupported_file_layout", "ZIP sem arquivos XLSX", 400)
    return xlsx_files


def parse_sinapi_package(path: Path, target_dir: Path) -> ParsedSinapiFile:
    xlsx_files = extract_xlsx_files(path, target_dir)
    reference = next((file for file in xlsx_files if "refer" in file.name.lower()), xlsx_files[0])
    return parse_consolidated_reference_xlsx(reference)


def parse_consolidated_reference_xlsx(path: Path) -> ParsedSinapiFile:
    workbook = load_workbook(path, read_only=True, data_only=True)
    ano, mes = detect_competencia(workbook)
    items_by_code: dict[str, dict] = {}

    for sheet_name, field_name in (
        ("ISE", "valor_sem_encargos"),
        ("ISD", "valor_onerado"),
        ("ICD", "valor_nao_onerado"),
    ):
        if sheet_name in workbook.sheetnames:
            parse_insumos_sheet(workbook[sheet_name], items_by_code, field_name)

    for sheet_name, field_name in (
        ("CSE", "valor_sem_encargos"),
        ("CSD", "valor_onerado"),
        ("CCD", "valor_nao_onerado"),
    ):
        if sheet_name in workbook.sheetnames:
            parse_composicoes_sheet(workbook[sheet_name], items_by_code, field_name)

    items = [
        ParsedItem(
            codigo=code,
            descricao=data["descricao"],
            tipo=data["tipo"],
            unidade=data.get("unidade"),
            classe=data.get("classe"),
            valor_sem_encargos=data.get("valor_sem_encargos"),
            valor_onerado=data.get("valor_onerado"),
            valor_nao_onerado=data.get("valor_nao_onerado"),
            raw_data=data.get("raw_data", {}),
        )
        for code, data in sorted(items_by_code.items())
    ]
    if not items:
        raise AppError("unsupported_file_layout", "Nenhum item SINAPI TO encontrado no XLSX", 400)
    return ParsedSinapiFile(
        uf="TO",
        ano=ano,
        mes=mes,
        source_layout="consolidated_xlsx_all_ufs",
        items=items,
        composicao_items=[],
    )


def detect_competencia(workbook) -> tuple[int, int]:
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(min_row=1, max_row=5, values_only=True):
            values = [str(value).strip() if value is not None else "" for value in row]
            for index, value in enumerate(values):
                if "Mês de Referência" in value or "Mes de Referencia" in value:
                    if index + 1 < len(values):
                        month_text = values[index + 1]
                        mes, ano = month_text.split("/")
                        return int(ano), int(mes)
    raise AppError("unsupported_file_layout", "Competencia nao encontrada no XLSX", 400)


def normalize_header(value) -> str:
    return str(value or "").replace("\n", " ").strip().upper()


def find_to_column(sheet, max_row: int = 12) -> tuple[int, int]:
    for row_index, row in enumerate(sheet.iter_rows(min_row=1, max_row=max_row, values_only=True), 1):
        for col_index, value in enumerate(row, 1):
            if normalize_header(value) == "TO":
                return row_index, col_index
    raise AppError("unsupported_file_layout", f"Coluna TO nao encontrada na aba {sheet.title}", 400)


def parse_insumos_sheet(sheet, items_by_code: dict[str, dict], field_name: str) -> None:
    _, to_col = find_to_column(sheet)
    for row in sheet.iter_rows(min_row=1, values_only=True):
        headers = [normalize_header(value) for value in row]
        if "CODIGO DO INSUMO" not in headers and "CÓDIGO DO INSUMO" not in headers:
            continue
        header_row = headers
        code_col = next(i for i, value in enumerate(header_row) if "CODIGO DO INSUMO" in value or "CÓDIGO DO INSUMO" in value)
        desc_col = next(i for i, value in enumerate(header_row) if "DESCRICAO DO INSUMO" in value or "DESCRIÇÃO DO INSUMO" in value)
        unit_col = next((i for i, value in enumerate(header_row) if value == "UNIDADE"), None)
        class_col = next((i for i, value in enumerate(header_row) if value == "CLASSIFICACAO" or value == "CLASSIFICAÇÃO"), None)
        start_row = row[0].row + 1 if hasattr(row[0], "row") else None
        break
    else:
        return

    # read_only values do not expose row numbers, so infer from observed CAIXA header contract.
    for data_row in sheet.iter_rows(min_row=10, values_only=True):
        if not data_row or normalize_header(data_row[code_col]) in {"CODIGO DO INSUMO", "CÓDIGO DO INSUMO"}:
            continue
        code = str(data_row[code_col] or "").strip()
        if not code:
            continue
        entry = items_by_code.setdefault(
            code,
            {
                "codigo": code,
                "descricao": str(data_row[desc_col] or "").strip(),
                "tipo": "INSUMO",
                "unidade": str(data_row[unit_col]).strip() if unit_col is not None and data_row[unit_col] else None,
                "classe": str(data_row[class_col]).strip() if class_col is not None and data_row[class_col] else None,
                "raw_data": {},
            },
        )
        entry[field_name] = parse_decimal(data_row[to_col - 1])


def parse_composicoes_sheet(sheet, items_by_code: dict[str, dict], field_name: str) -> None:
    _, to_col = find_to_column(sheet)
    for data_row in sheet.iter_rows(min_row=10, values_only=True):
        headers = [normalize_header(value) for value in data_row]
        if "CODIGO DA COMPOSICAO" in headers or "CÓDIGO DA COMPOSIÇÃO" in headers:
            continue
        if len(data_row) < 4:
            continue
        code = str(data_row[1] or "").strip()
        if not code:
            continue
        entry = items_by_code.setdefault(
            code,
            {
                "codigo": code,
                "descricao": str(data_row[2] or "").strip(),
                "tipo": "COMPOSICAO",
                "unidade": str(data_row[3]).strip() if data_row[3] else None,
                "classe": str(data_row[0]).strip() if data_row[0] else None,
                "raw_data": {},
            },
        )
        entry[field_name] = parse_decimal(data_row[to_col - 1])
